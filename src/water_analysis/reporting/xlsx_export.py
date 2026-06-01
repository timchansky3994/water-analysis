"""CSV and human-readable XLSX table export helpers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

COLUMN_LABELS_RU: dict[str, str] = {
    "scope_name": "Сценарий анализа",
    "scope_id": "Идентификатор среза",
    "scope_label": "Описание среза",
    "target": "Целевой показатель",
    "feature": "Сопутствующий показатель",
    "method": "Метод",
    "corr": "Коэффициент корреляции",
    "n_shared": "Совместных измерений",
    "p_value": "p-value",
    "status": "Статус",
    "Indicator": "Показатель",
    "record_count": "Число записей",
    "sample_event_count": "Фактов взятия проб",
    "numeric_observation_count": "Число числовых значений",
    "missing_numeric_count": "Число пропусков",
    "missing_ratio": "Доля пропусков",
    "censored_count": "Цензурированных значений",
    "censored_ratio": "Доля цензурированных значений",
    "total_rows": "Всего строк",
    "missing_count": "Число пропусков",
    "PointType_Code": "Тип точки",
    "unique_points": "Уникальных точек",
    "unique_oktmo": "Уникальных ОКТМО",
    "indicator_count": "Число показателей",
    "n_observations": "Число наблюдений",
    "n_unique": "Уникальных значений",
    "dominant_share": "Доля доминирующего значения",
    "indicator_x": "Показатель 1",
    "indicator_y": "Показатель 2",
    "sample_point_rows": "Строк в формате точка-дата",
    "target_observation_count": "Наблюдений целевого показателя",
    "target_missing_ratio": "Доля пропусков целевого показателя",
    "target_censored_ratio": "Доля цензурирования целевого показателя",
    "eligible_predictor_count": "Доступных предикторов",
    "max_shared_samples": "Максимум совместных измерений",
    "issue_codes": "Коды причин",
    "issue_messages": "Пояснения",
    "model_name": "Модель",
    "model_family": "Тип модели",
    "is_baseline": "Базовая модель",
    "selected_feature_count": "Число выбранных признаков",
    "selected_features": "Выбранные признаки",
    "beats_best_baseline": "Лучше базовой модели",
    "comparison_note": "Итог сравнения",
    "notes": "Примечания",
    "holdout_mae": "MAE на проверке",
    "holdout_rmse": "RMSE на проверке",
    "holdout_r2": "R2 на проверке",
    "holdout_smape": "SMAPE на проверке",
    "backtest_mae": "MAE backtest",
    "backtest_rmse": "RMSE backtest",
    "backtest_r2": "R2 backtest",
    "backtest_smape": "SMAPE backtest",
    "combined_score": "Комбинированный скор",
    "stability_ratio": "Стабильность (backtest/holdout)",
    "combined_score_used_fallback": "Использован fallback на holdout",
    "top_interpretation": "Важные признаки",
    "SampleDate": "Дата отбора",
    "FullPointCode": "Полный код точки",
    "OKTMO": "ОКТМО",
    "PointNumber": "Номер точки",
    "actual": "Фактическое значение",
    "predicted": "Расчетное значение",
    "residual": "Остаток модели",
    "target_corr": "Корреляция с целевым показателем",
    "target_corr_abs": "Модуль корреляции",
    "target_corr_p_value": "p-value корреляции",
    "prediction_status": "Статус расчетной оценки",
    "predicted_value": "Расчетное значение",
    "lower_bound": "Нижняя граница",
    "upper_bound": "Верхняя граница",
    "model_package": "Пакет модели",
    "feature_names_used": "Использованные признаки",
    "features_total": "Всего признаков модели",
    "features_observed": "Доступных признаков",
    "feature_coverage": "Покрытие признаков",
    "missing_features": "Отсутствующие признаки",
    "warnings": "Предупреждения",
    "Value_Approx": "Числовое значение",
    "ValueSource": "Источник значения",
    "IsEstimated": "Расчетное значение",
    "ModelPackage": "Пакет модели",
    "OriginalValuePresent": "Было исходное лабораторное значение",
    "level": "Уровень диагностики",
    "reason": "Причина",
    "detail": "Детали",
    "included": "Выбрана",
    "exclusion_reason": "Причина невыбора",
    "selection_score": "Оценка отбора",
    "is_forced": "Задана специалистом",
    "selection_mode": "Режим отбора",
    "forced_features": "Заданные специалистом признаки",
}

SHEET_NAMES_RU: dict[str, str] = {
    "profile_summary": "Паспорт данных",
    "indicator_observations": "Наблюдения по показателям",
    "missingness": "Пропуски",
    "point_type_coverage": "Типы точек",
    "constant_series": "Постоянные ряды",
    "cooccurrence": "Совместные измерения",
    "correlation_results": "Корреляции",
    "correlation_diagnostics": "Диагностика корреляций",
    "readiness": "Пригодность",
    "comparison_summary": "Сравнение моделей",
    "holdout_predictions": "Проверочные прогнозы",
    "backtest_summary": "Backtest",
    "feature_selection_candidates": "Отбор признаков",
    "interpretability": "Интерпретация",
    "predictions": "Расчетные оценки",
    "estimated_values_long": "Long format",
    "inference_diagnostics": "Диагностика",
}


def humanize_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Return a copy whose columns carry the Russian labels used in XLSX exports.

    Only the column headers are translated; cell values are untouched. The UI
    uses this so on-screen tables match the downloadable XLSX files exactly.
    """
    exported = dataframe.copy()
    exported.columns = [COLUMN_LABELS_RU.get(str(column), str(column)) for column in exported.columns]
    return exported


# Backwards-compatible internal alias.
_xlsx_dataframe = humanize_columns


def _apply_formatting(workbook) -> None:
    """Apply basic Excel formatting for readability to an open workbook."""
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:100]]
            width = min(max(max((len(value) for value in values), default=0) + 2, 12), 60)
            worksheet.column_dimensions[column_letter].width = width


def _format_workbook(path: Path) -> None:
    """Apply basic Excel formatting to a workbook saved on disk."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    _apply_formatting(workbook)
    workbook.save(path)


def dataframe_to_xlsx_bytes(dataframe: pd.DataFrame, *, table_name: str) -> bytes:
    """Render a dataframe as a formatted human-readable XLSX file in memory.

    Used by the Streamlit UI to offer XLSX downloads for tables that are
    computed on the fly and never written to disk.
    """
    import io

    from openpyxl import load_workbook

    sheet_name = SHEET_NAMES_RU.get(table_name, table_name[:31])[:31]
    raw_buffer = io.BytesIO()
    with pd.ExcelWriter(raw_buffer, engine="openpyxl") as writer:
        humanize_columns(dataframe).to_excel(writer, index=False, sheet_name=sheet_name)
    raw_buffer.seek(0)
    workbook = load_workbook(raw_buffer)
    _apply_formatting(workbook)
    formatted_buffer = io.BytesIO()
    workbook.save(formatted_buffer)
    return formatted_buffer.getvalue()


def save_dataframe_xlsx(dataframe: pd.DataFrame, path: str | Path, *, table_name: str) -> Path:
    """Save a dataframe as a human-readable XLSX file."""
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_name = SHEET_NAMES_RU.get(table_name, table_name[:31])[:31]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _xlsx_dataframe(dataframe).to_excel(writer, index=False, sheet_name=sheet_name)
    _format_workbook(output_path)
    return output_path


def save_dataframe_csv_and_xlsx(dataframe: pd.DataFrame, csv_path: str | Path, *, table_name: str) -> dict[str, Path]:
    """Save technical CSV and adjacent human-readable XLSX versions."""
    output_path = Path(csv_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(output_path, index=False, encoding="utf-8-sig")
    xlsx_path = output_path.with_suffix(".xlsx")
    save_dataframe_xlsx(dataframe, xlsx_path, table_name=table_name)
    return {"csv": output_path, "xlsx": xlsx_path}
